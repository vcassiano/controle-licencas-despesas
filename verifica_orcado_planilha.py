import openpyxl

# Caminho do arquivo
arquivo = "Orcamento_TI_CAPEX_OPEX_2026-v1 (1).xlsx"

wb = openpyxl.load_workbook(arquivo, data_only=True)

somas = {}
for aba in ["Base_OPEX", "Base_CAPEX"]:
    if aba in wb.sheetnames:
        ws = wb[aba]
        total = 0
        for row in ws.iter_rows(min_row=4):
            ano = row[0].value
            if not ano:
                break
            # Soma todos os meses orçados (colunas N, R, V, Z, AD, AH, AL, AP, AT, AX, BB, BF)
            meses_cols = [13, 17, 21, 25, 29, 33, 37, 41, 45, 49, 53, 57]
            soma_item = sum(float(row[col].value or 0) for col in meses_cols)
            total += soma_item
        somas[aba] = total

print("OPEX Orçado:", somas.get("Base_OPEX", 0))
print("CAPEX Orçado:", somas.get("Base_CAPEX", 0))
print("Total Orçado:", somas.get("Base_OPEX", 0) + somas.get("Base_CAPEX", 0))
